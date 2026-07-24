"""
reports.py
==========
Transforma os resultados do analytics.py num ficheiro Excel.
"""

import openpyxl
from utils import ensure_dir
from config import OUTPUT_DIR


def export_dict_to_sheet(workbook, nome_folha, dicionario, nome_coluna_chave, nome_coluna_valor):
    """Escreve um dicionário {chave: valor} como uma tabela de 2 colunas numa nova folha."""
    folha = workbook.create_sheet(nome_folha)
    folha.cell(row=1, column=1, value=nome_coluna_chave)
    folha.cell(row=1, column=2, value=nome_coluna_valor)

    linha = 2
    for chave, valor in dicionario.items():
        folha.cell(row=linha, column=1, value=str(chave))
        folha.cell(row=linha, column=2, value=valor)
        linha += 1


def export_report(eventos_por_dia, overtime_por_dia, horas_por_cliente, horas_por_tipo):
    """Cria o ficheiro Excel com uma folha para cada análise."""
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # remove a folha vazia que vem por defeito

    export_dict_to_sheet(workbook, "Horas por Dia", eventos_por_dia, "Data", "Horas")
    export_dict_to_sheet(workbook, "Overtime por Dia", overtime_por_dia, "Data", "Overtime")
    export_dict_to_sheet(workbook, "Horas por Cliente", horas_por_cliente, "Cliente", "Horas")
    export_dict_to_sheet(workbook, "Horas por Tipo", horas_por_tipo, "Tipo", "Horas")

    pasta = ensure_dir(OUTPUT_DIR)
    caminho = f"{pasta}/relatorio.xlsx"
    workbook.save(caminho)
    return caminho